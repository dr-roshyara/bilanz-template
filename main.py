from tkinter import *
from tkinter import ttk
# from glob import glob
from msilib.schema import Font
# from PIL import ImageTk,Image 
# import PIL.Image 
from tkinter import filedialog,messagebox
# from turtle import color
from openpyxl  import Workbook, load_workbook
from openpyxl.utils import get_column_letter
# import json
from datetime import datetime
import locale
from openpyxl.styles.colors import Color
from openpyxl.styles import Font,Side,Border,PatternFill
from openpyxl.formula.translate import Translator 

#set local parameter 
locale.setlocale(locale.LC_ALL, 'deu_deu')


def Slide(Progress_Bar):
    import time
    Progress_Bar['value']=20
    window.update_idletasks()
    time.sleep(1)
    Progress_Bar['value']=40
    window.update_idletasks()
    time.sleep(1)
    Progress_Bar['value']=60
    window.update_idletasks()
    time.sleep(0.5)
    Progress_Bar['value']=80
    time.sleep(0.5)
    Progress_Bar['value']=100



def create_bilanz():
    
    file_info   =get_file_info();
    firmen_info =get_firmen_info(file_info[0]);
    original_filename=file_info[0]+file_info[1];
    #write bilanz
    filename    =file_info[0]+"bilanz_"+file_info[1];
    show_image_icon();
    # show_image_icon();
    # my_img      = PhotoImage(file = "downloadicon.gif") 
    # imageGrid   = Label(window,  image=my_img )
    # imageGrid.grid(row=1,column=2) 
    main_wb     =read_main_file(original_filename);    #
    create_bilanz_file(filename, main_wb, firmen_info,file_info[1])
    #show dialogue 
    inform_success_bilanz_work();
    window.destroy() 
#
def show_image_icon():
    #function to read the file 
    Progress_Bar=ttk.Progressbar(window,orient=HORIZONTAL,length=250,mode='determinate')
    Progress_Bar.pack()
    Slide(Progress_Bar);

    # canvas = Canvas(window, width = 500, height = 500)
    # canvas.pack()
    #GIF in my_image variable
    #Give the entire file address along with the file name and gif extension
    #Use \\ in the address
    #The image given by me is C:\\UserAdmin\\Device\\Desktop2\\canyon.gif
    # my_image = PhotoImage(file="./assets/downloadicon.gif")
    # canvas.create_image(0, 0, anchor = NW, image=my_image)

    # fp = open("./assets/downloadicon.gif","rb")
    # render = ImageTk.PhotoImage(fp)
    # img = PIL.Image.open(fp)
    # img.show()
    # load = Image.open("./assets/downloadicon.gif")
    # render = ImageTk.PhotoImage(load, format="gif -index 2")
    # img = Label(window, image=render)
    # img.image = render
    # img.place(x=20, y=10)
    # image = Image.open('./assets/downloadicon.gif')
    # python_image = ImageTk.PhotoImage(image)
    # canvas = Canvas(window, width = 200, height = 200)      
    # canvas.pack()      
    # img = PhotoImage(file="downloadicon.gif")      
    # canvas.create_image(10,10, anchor=NW, image=img)   
    # self.img = PhotoImage(Image.open("downloadicon.gif")) 
    # self.canvas.create_image(20,20, anchor=NW, image=self.img)    
    # self.canvas.image = self.img   
    return ;
#
def inform_success_bilanz_work():
    messagebox.showinfo ("Bilanz  erstellt!", 
        "Bilanz Datein erfolgreich erstellt");


    return ;
#    
def get_firmen_info(file_dir):
    _file =file_dir+"firmen_info.xlsx";
    wb =load_workbook(_file)
    ws =wb.active
    total_rows =len(ws["A"])+1;
    keyList=[];
    valueList=[];
    for row in range(2, total_rows):
        # for col in range(1,2)
        key_char    =get_column_letter(1);
        value_char  =get_column_letter(2);        
        _key        =str(ws[key_char+str(row)].value);
        _value      =str(ws[value_char+str(row)].value);
        keyList.append(_key);
        valueList.append(_value);
    firmenname_index    =keyList.index('firmenname')
    bilanz_index        =keyList.index('Bilanzdate')    
    firmenname          =valueList[firmenname_index]
    bilanzDate          =valueList[bilanz_index];
    bilanzDate          = datetime.strptime(bilanzDate,"%Y-%m-%d  %H:%M:%S")
    # print(bilanzDate.strftime("%d. %B %Y"));
    return [firmenname,bilanzDate];




def get_file_info():
    global file_path;
    file_path =filedialog.askopenfilename();
    _arr      =file_path.split("/");
    filename =_arr[-1];
    file_dir =_arr[0:(len(_arr)-1)]
    # print(file_dir);
    real_path ="/".join(file_dir)+"/"; 
    # print(real_path);
    return [real_path, filename];
"""
  @param: filename 
  @return: workbook 
  #This function reads a excel file and reads a workbook 
  #   
"""

def read_main_file(filename):
    wb          =load_workbook(filename, data_only=True)
    #if you dont use data_only=True , then you will read the formula also.
    return wb;

#create Bilanz  file 
#######################################################################
def create_bilanz_file(filename, main_wb, firmen_info,original_filename):
    #Select the main file to read the data and create bilanz   
    sheetName       ="XX05"
    # ws            =main_wb.sheetName
    sheet_names     =main_wb.sheetnames;
    _xx05           = sheet_names.index('XX05');
    main_wb.active  =main_wb[sheet_names[_xx05]];
    # main_wb.active =main_wb['xx05'];
    main_ws         =main_wb.active;

    #Create new workbook for Bilanz 
    wb =Workbook();
    wb.title    = "Bilanz"
    new_ws =wb.active;
    new_ws.title="Bilanz"
    # sheetnames  =wb.sheetnames
    # for sheet in sheetnames:
    #   wb.remove(sheet);

    #create a new sheet named Bilanz
    # create header;
    new_ws["A2"].value  =firmen_info[0];
    new_ws["A3"].value  ="Bilanz zum "+str( firmen_info[1].strftime("%d. %B %Y"));      
    new_ws["A2"].font   =Font(bold=True, size=18,color="0E0E0E");
    new_ws["A3"].font   =Font(bold=True,size=16, color="0E0E0E")
    new_ws["A3"].border = Border(bottom=Side(border_style="thick", color='0E0E0E') );
    new_ws["Q4"].value  ="Anlage 1";
    ###################################################################################
    # #start creating bilanz template
    new_ws["A5"].value  ="Mapping ID"
    new_ws["B5"].value  ="Aktiva"
    #
    new_ws["B5"].font   = Font(bold=True,size=12, color="0E0E0E")   
    new_ws["E5"].value  =firmen_info[1].strftime("%d.%m.%Y");
    new_ws.column_dimensions['E'].width = float(18)
    #make yellow color background
    new_ws["F5"].value  ="EY Ref"
    new_ws["F5"].font   = Font(bold=True,size=12, color="FF0000")   
    new_ws["J5"].value  ="Mapping ID"
    new_ws["K5"].value  ="Passiva";
    new_ws["K5"].font   = Font(bold=True,size=12, color="0E0E0E")   
    new_ws["Q5"].value  =firmen_info[1].strftime("%d.%m.%Y");
    new_ws["S5"].value  ="davon-Vermerke" 
    ###################################################################################
    #set the column width 
    # column C large 
    new_ws.column_dimensions['A'].width = float(12)   
    new_ws.column_dimensions['J'].width = float(12)   
    new_ws.column_dimensions['C'].width = float(60)   
    new_ws.column_dimensions['G'].width = float(31)
    new_ws.column_dimensions['L'].width = float(65)
    new_ws.column_dimensions['M'].width = float(18)
    new_ws.column_dimensions['O'].width = float(35)
    new_ws.column_dimensions['P'].width = float(30)
    new_ws.column_dimensions['Q'].width = float(18)
    new_ws.column_dimensions['R'].width = float(20)
    new_ws.column_dimensions['S'].width = float(20)
    ###################################################
    #set column color: 
    #make color for column F and G: ecf3ea
    for col_range in range(5, 57):
        new_ws['F'+str(col_range)].fill = PatternFill(fill_type="lightVertical", start_color='E6E9B9', end_color='E6E9B9')
        new_ws['G'+str(col_range)].fill = PatternFill(fill_type="lightVertical", start_color='E6E9B9', end_color='E6E9B9')
    #make color for column O and P 
    for col_range in range(5, 57):
        new_ws['O'+str(col_range)].fill = PatternFill(fill_type="lightVertical", start_color='E6E9B9', end_color='E6E9B9')
        new_ws['P'+str(col_range)].fill = PatternFill(fill_type="lightVertical", start_color='E6E9B9', end_color='E6E9B9')
    #make color for Column R :
    for col_range in range(5, 57):
        new_ws['R'+str(col_range)].fill = PatternFill(fill_type="lightVertical", start_color='E6E9B9', end_color='E6E9B9')
    
    ###################################################################################
    #Start of create mappping id first column
    new_ws["A14"].value="A.A.I.2."
    new_ws["A15"].value="A.A.I.3."
    #
    new_ws["A23"].value="A.A.II.1."
    new_ws["A24"].value="A.A.II.2."
    new_ws["A25"].value="A.A.II.3."
    new_ws["A26"].value="A.A.II.4."
    #
    new_ws["A37"].value="A.B.II.2."
    new_ws["A38"].value="A.B.II.3."
    new_ws["A39"].value="A.B.II.5."
    #
    new_ws["A43"].value="A.B.IV."
    new_ws["A48"].value="A.C"
    ###################################################################################
    #coloumn B
    new_ws["B9"].value  ="A."
    new_ws["B9"].font   =Font(bold=True,size=13, color="0E0E0E")
    new_ws["B11"].value  ="I."
    new_ws["B11"].font   =Font(bold=True,size=13, color="0E0E0E")
    new_ws["B13"].value  ="1."
    new_ws["B15"].value  ="2."
    #
    new_ws["B20"].value  ="II."
    new_ws["B20"].font   =Font(bold=True,size=13, color="0E0E0E")
    #
    new_ws["B22"].value  ="1."
    new_ws["B24"].value  ="2."
    new_ws["B25"].value  ="3."
    new_ws["B26"].value  ="4."
    #
    new_ws["B33"].value  ="B."
    new_ws["B33"].font   =Font(bold=True,size=13, color="0E0E0E")
    #
    new_ws["B35"].value  ="I."
    new_ws["B35"].font   =Font(bold=True,size=13, color="0E0E0E")
    #
    new_ws["B37"].value  ="1."
    new_ws["B38"].value  ="2."
    new_ws["B39"].value  ="3."
    #
    new_ws["B42"].value  ="II."
    new_ws["B42"].font   =Font(bold=True,size=13, color="0E0E0E")
   #
    new_ws["B48"].value  ="C."
    new_ws["B48"].font   =Font(bold=True,size=13, color="0E0E0E")
     
    
    ###################################################################################
    #column C 
    new_ws["C9"].value   ="Anlagevermögen"
    new_ws["C9"].font    =Font(bold=True,size=13, color="0E0E0E")
    new_ws["C11"].value  ="Immaterielle Vermögensgegenstände"
    new_ws["C11"].font   =Font(bold=True,size=13, color="0E0E0E")
    new_ws["C13"].value  ="Entgeltlich erworbene Konzessionen, gewerbliche Schutzrechte und ähnliche Rechte"
    new_ws["C14"].value  ="und Werte sowie Lizenzen an solchen Rechten und Werten"
    new_ws["C15"].value  ="Geschäfts- oder Firmenwert"
    #
    new_ws["C33"].value  ="Umlaufvermögen"
    new_ws["C33"].font   =Font(bold=True,size=13, color="0E0E0E")
    #
    new_ws["C33"].value  ="Forderungen und sonstige Vermögensgegenstände"
    new_ws["C33"].font   =Font(bold=True,size=13, color="0E0E0E")
    #
    new_ws["C37"].value  ="Forderungen aus Lieferungen und Leistungen"
    new_ws["C38"].value  ="Forderungen gegen verbundene Unternehmen"
    new_ws["C39"].value  ="Sonstige Vermögensgegenstände"
    #
    new_ws["C42"].value  ="Kassenbestand, Bundesbankguthaben,"
    new_ws["C42"].font   =Font(bold=True,size=13, color="0E0E0E")
    #
    new_ws["C43"].value  ="Guthaben bei Kreditinstituten"
    new_ws["C43"].font   =Font(bold=True,size=13, color="0E0E0E")
    #
    new_ws["C48"].value  ="Rechnungsabgrenzungsposten"
    new_ws["C48"].font   =Font(bold=True,size=13, color="0E0E0E")
   
    
    
    ###################################################################################
    #Column D 
    # Read the data from main worksheet: xx05
    mapping_ids =[];
    total_amount=[];
    for _col in main_ws["A"]:
        _colValue=str(_col.value);
        mapping_ids.append(_colValue.strip());

    for _col in main_ws["S"]:
        _colValue =_col.internal_value;
        total_amount.append(_colValue);
    # print(mapping_ids)
    _search_string ='A.A.I.2.';
    indices = [i for i, x in enumerate(mapping_ids) if x ==_search_string];
    # https://stackoverflow.com/questions/6294179/how-to-find-all-occurrences-of-an-element-in-a-list
    #get the sum
    # https://stackoverflow.com/questions/1012185/in-python-how-do-i-index-a-list-with-another-list    
    sub_total   = [total_amount[i] for i in indices];
    new_ws["D14"].value=sum(sub_total);
    #for new_wb["D15"]
    indices=[];
    sub_total1=[];
    _filename =original_filename
    _formula_string = '=SUMIF([';
    _formula_string += original_filename;
    _formula_string += "]XX05!$A:$A;Bilanz!$A$15;["
    _formula_string += original_filename;
    _formula_string += "]XX05!$S:$S)";
    # print(_formula_string);    
    # Get Sum for the mapping id :A.A.I.3
    _search_string ="A.A.I.3.";
    indices = [i for i, x in enumerate(mapping_ids) if x ==_search_string];
    sub_total1   = [total_amount[i] for i in indices];
    new_ws["D15"].value=sum(sub_total1);
    new_ws["D15"].border = Border(bottom=Side(border_style="thin", color='0E0E0E') );
    # new_ws["E19"].value =sum(sub_total)+sum(sub_total1);
    new_ws["E17"].value ="=$D$14+$D$15";
    new_ws["E18"].value  ="|--------- @ ---------|"
    new_ws["E18"].font   =Font(bold=True,size=12, color="DB1203")
    #################################################################
    #Column F 
    new_ws["F14"].value="L-Lead"
    new_ws["F14"].font   =Font(bold=True,size=12, color="DB1203")
    #
    new_ws["F15"].value  ="L-Lead"
    new_ws["F15"].font   =Font(bold=True,size=12, color="DB1203")
   #
    new_ws["F17"].value  ="L-Lead"
    new_ws["F17"].font   =Font(bold=True,size=12, color="DB1203")
    ###############################################################
    #column G 
    new_ws["G15"].value  ="A-04 Lagebericht"
    new_ws["G15"].font   =Font(bold=True,size=12, color="DB1203")
    #
    new_ws["G38"].value  ="A-03 Anhang; A-04 Lagebericht"
    new_ws["G38"].font   =Font(bold=True,size=12, color="DB1203")

    #
    new_ws["G43"].value  ="A-04 Lagebericht"
    new_ws["G43"].font   =Font(bold=True,size=12, color="DB1203")
    ###############################################################
    #column J
    new_ws["J10"].value  ="P.A.I."
    new_ws["J12"].value  ="P.A.IV."
    new_ws["J14"].value  ="P.A.V."
    #
    new_ws["J21"].value  ="P.B.1."
    new_ws["J22"].value  ="P.B.2."
    new_ws["J23"].value  ="P.B.3."
    #
    new_ws["J30"].value  ="P.C.4."
    new_ws["J31"].value  ="P.C.5."
    new_ws["J32"].value  ="P.C.6."
    ###############################################################
    #Column K 
    new_ws["K8"].value  ="A."
    new_ws["K8"].font   =Font(bold=True,size=13, color="0E0E0E")
    #
    new_ws["K10"].value  ="I."
    new_ws["K10"].font   =Font(bold=True,size=13, color="0E0E0E")
    #
    new_ws["K12"].value  ="II."
    new_ws["K12"].font   =Font(bold=True,size=13, color="0E0E0E")
    #
    new_ws["K14"].value  ="III."
    new_ws["K14"].font   =Font(bold=True,size=13, color="0E0E0E")
    #
    new_ws["K19"].value  ="B."
    new_ws["K19"].font   =Font(bold=True,size=13, color="0E0E0E")
    #
    new_ws["K21"].value  ="1."
    new_ws["K22"].value  ="2."
    new_ws["K23"].value  ="3."
     #
    new_ws["K28"].value  ="C."
    new_ws["K28"].font   =Font(bold=True,size=13, color="0E0E0E")
    #
    new_ws["K30"].value  ="1."
    new_ws["K31"].value  ="2."
    new_ws["K32"].value  ="3."
    #################################################################
    #Column L
    new_ws["L8"].value  ="Eigenkapital"
    new_ws["L8"].font   =Font(bold=True,size=13, color="0E0E0E")
    #
    new_ws["L10"].value  ="Gezeichnetes Kapital"
    new_ws["L10"].font   =Font(bold=True,size=13, color="0E0E0E")
    #
    new_ws["L12"].value  ="Gewinnvortrag"
    new_ws["L12"].font   =Font(bold=True,size=13, color="0E0E0E")
    #
    new_ws["L14"].value  ="Jahresüberschuss"
    new_ws["L14"].font   =Font(bold=True,size=13, color="0E0E0E")
    #
    new_ws["L19"].value  ="Rückstellungen"
    new_ws["L19"].font   =Font(bold=True,size=13, color="0E0E0E")
    #
    new_ws["L21"].value  ="Rückstellungen für Pensionen und ähnliche Verpflichtungen"
    new_ws["L22"].value  ="Steuerrückstellungen"
    new_ws["L23"].value  ="Sonstige Rückstellungen"
    #
    new_ws["L28"].value  ="Verbindlichkeiten"
    new_ws["L28"].font   =Font(bold=True,size=13, color="0E0E0E")
   #
    new_ws["L30"].value  ="Verbindlichkeiten aus Lieferungen und Leistungen"
    new_ws["L31"].value  ="Verbindlichkeiten gegenüber verbundenen Unternehmen"
    new_ws["L32"].value  ="Sonstige Verbindlichkeiten"
    new_ws["L33"].value  ="davon aus Steuern EUR 192.422,35 (Vj. TEUR 98)"
    new_ws["L34"].value  ="davon im Rahmen der sozialen Sicherheit EUR 171.203,70 (Vj. TEUR 167)"
    #
    new_ws["L54"].value  ="KONTROLLE"
    new_ws["L55"].value  ="EK-Quoute"
    ##################################################################    
    #column M 
    #P.B.1.
    _cell               =new_ws["M21"]
    sub_total           =[];
    _search_string      =new_ws["J21"].value;
    indices             = [i for i, x in enumerate(mapping_ids) if x ==_search_string];
    sub_total           = [total_amount[i] for i in indices];
    _cell.value         =sum(sub_total);
    _cell.value         =-1*round(sum(sub_total),2);
    #M22
    _cell               =new_ws["M22"]
    sub_total           =[];
    _search_string      =new_ws["J22"].value;
    indices             = [i for i, x in enumerate(mapping_ids) if x ==_search_string];
    sub_total           = [total_amount[i] for i in indices];
    _cell.value         =sum(sub_total);
    _cell.value         =-1*round(sum(sub_total),2);
    #M23
    _cell               =new_ws["M23"]
    sub_total           =[];
    _search_string      =new_ws["J23"].value;
    indices             = [i for i, x in enumerate(mapping_ids) if x ==_search_string];
    sub_total           = [total_amount[i] for i in indices];
    _cell.value         =sum(sub_total);
    _cell.value         =-1*round(sum(sub_total),2);
    new_ws["M23"].border = Border(bottom=Side(border_style="thin", color='0E0E0E') );
    #M30
    _cell               =new_ws["M30"]
    sub_total           =[];
    _search_string      =new_ws["J30"].value;
    indices             = [i for i, x in enumerate(mapping_ids) if x ==_search_string];
    sub_total           = [total_amount[i] for i in indices];
    _cell.value         =sum(sub_total);
    _cell.value         =-1*round(sum(sub_total),2);
    #M31
    _cell               =new_ws["M31"]
    sub_total           =[];
    _search_string      =new_ws["J31"].value;
    indices             = [i for i, x in enumerate(mapping_ids) if x ==_search_string];
    sub_total           = [total_amount[i] for i in indices];
    _cell.value         =sum(sub_total);
    _cell.value         =-1*round(sum(sub_total),2);
    
    #M32
    _cell               =new_ws["M32"]
    sub_total           =[];
    _search_string      =new_ws["J32"].value;
    indices             = [i for i, x in enumerate(mapping_ids) if x ==_search_string];
    sub_total           = [total_amount[i] for i in indices];
    _cell.value         =-1*round(sum(sub_total),2);
    # _cell.number_format = '#.##0,00'
    #
    new_ws["M33"].value  ="PY";
    new_ws["M33"].font   =Font(bold=True,size=12, color="1B0CF5")
    #
    new_ws["M34"].value  ="PY";
    new_ws["M34"].font   =Font(bold=True,size=12, color="1B0CF5")
    ##################################################################
    # Column O:
    new_ws["O6"].value  ="EY  Ref"
    new_ws["O6"].font   =Font(bold=True,size=12, color="DB1203")
    
    #
    new_ws["O10"].value  ="A-03 Anhang"
    new_ws["O10"].font   =Font(bold=True,size=12, color="DB1203")
    #
    new_ws["O12"].value  ="A-03 Anhang"
    new_ws["O12"].font   =Font(bold=True,size=12, color="DB1203")
   #
    new_ws["O16"].value  ="A-04 Lagebericht"
    new_ws["O16"].font   =Font(bold=True,size=12, color="DB1203")
   #
    new_ws["O21"].value  ="A-03 Anhang; A-04 Lagebericht"
    new_ws["O21"].font   =Font(bold=True,size=12, color="DB1203")
    #
    new_ws["O31"].value  ="A-03 Anhang; A-04 Lagebericht"
    new_ws["O31"].font   =Font(bold=True,size=12, color="DB1203")
     #
    new_ws["O33"].value  ="tab XX05"
    new_ws["O33"].font   =Font(bold=True,size=12, color="DB1203")
    #
    new_ws["O34"].value  ="tab XX05"
    new_ws["O34"].font   =Font(bold=True,size=12, color="DB1203")
    ##################################################################
    #Column P: 
    new_ws["R6"].value  ="EY Ref"
    new_ws["R6"].font   =Font(bold=True,size=12, color="DB1203")
    ##################################################################
    #Column R: 
    new_ws["P10"].value  ="T-Lead"
    new_ws["P10"].font   =Font(bold=True,size=12, color="DB1203")
    #
    new_ws["P12"].value  ="T-Lead"
    new_ws["P12"].font   =Font(bold=True,size=12, color="DB1203")
    #
    new_ws["P14"].value  ="T-Lead"
    new_ws["P14"].font   =Font(bold=True,size=12, color="DB1203")
    #
    new_ws["P16"].value  ="T-Lead"
    new_ws["P16"].font   =Font(bold=True,size=12, color="DB1203")
    #
    new_ws["P21"].value  ='tab "EY Lead Sheets--> RST"'
    new_ws["P21"].font   =Font(bold=True,size=12, color="DB1203")
    #
    new_ws["P22"].value  ='tab "EY Lead Sheets--> RST"'
    new_ws["P22"].font   =Font(bold=True,size=12, color="DB1203")
    #
    new_ws["P23"].value  ='tab "EY Lead Sheets--> RST"'
    new_ws["P23"].font   =Font(bold=True,size=12, color="DB1203")
    #
    new_ws["P25"].value  ='P-Lead'
    new_ws["P25"].font   =Font(bold=True,size=12, color="DB1203")
    #
    new_ws["P30"].value  ='N-Lead'
    new_ws["P30"].font   =Font(bold=True,size=12, color="DB1203")
    #
    new_ws["P31"].value  ='I2-Lead'
    new_ws["P31"].font   =Font(bold=True,size=12, color="DB1203")
    #
    new_ws["P32"].value  ='P3-Lead'
    new_ws["P32"].font   =Font(bold=True,size=12, color="DB1203")
    
    ##################################################################
    #Column R: 
    # new_ws["R12"].value  =+1*new_ws["N12"].value-new_ws["Q12"].value
    new_ws["R12"].value  =round(0.00,2)
    new_ws["R12"].number_format ="0.00"
    #
    new_ws["R16"].value  ="A-04 Lagebericht"
    new_ws["R16"].font   =Font(bold=True,size=12, color="DB1203")
    #
    new_ws["R56"].value  ="A-04 Lagebericht"
    new_ws["R56"].font   =Font(bold=True,size=12, color="DB1203")

    ##################################################################
    #finally save the file
    wb.save(filename)
    ##################################################################
    
#get the main window 
#######################################################################
def get_window():
    global window;
    window = Tk()
    window.title("Bilanz Erstellung");
    window.config(bg='#D8FFD9')
    window.geometry("400x400")      
    #create a button to upload the file 
    button_get_path = Button(window, text=" Bitte wählen Sie Ihre Jahresabschluss Datei!  ", command=create_bilanz)
    button_get_path.place(x=20, y=20)    
    button_get_path.pack(expand=True)
    window.eval('tk::PlaceWindow . center') 
    window.mainloop()
   

get_window()
window.destroy() 
